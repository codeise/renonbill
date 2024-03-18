from itertools import islice
import re
from collections import OrderedDict
from openpyxl import load_workbook


def clean(s):
    if s is None:
        return None
    if isinstance(s, int):
        s = "p" + str(s)
    s = s.lower()
    # Remove invalid characters
    s = re.sub('[^0-9a-zA-Z_]', '', s)
    # Remove leading characters until we find a letter or underscore
    s = re.sub('^[^a-zA-Z_]+', '', s)
    return s


def header_name(s, key_replacements):
    if key_replacements and (s in key_replacements):
        s = key_replacements[s]
    s = clean(s)
    if isinstance(s, str):
        s = s.strip()
    return s


def value(s, value_replacements, nulls_to_empty):
    if s is None and nulls_to_empty:
        s = ''
    if value_replacements and (s in value_replacements):
        s = value_replacements[s]
    if isinstance(s, str):
        s = s.strip()
    return s


def all_empty(row):
    all_empty_values = True
    for row_key in row:
        if row[row_key] is not None and row[row_key] != "":
            all_empty_values = False
            break
    return all_empty_values


def import_standard_table(work_sheet, config):
    json_list = []
    row_start = 1
    row_end = None
    column_start = 0
    column_end = None
    custom_headers = None
    if config.get("range"):
        range_v = config.get("range")
        row_start = range_v.get("row_start")
        row_end = range_v.get("row_end")
        column_start = range_v.get("column_start", 1) - 1
        column_end = range_v.get("column_end")
        # if column_end:
        #     column_end -= 1
    if config.get("custom_headers"):
        headers = config.get("custom_headers")
        custom_headers = True
        row_start -= 1
    else:
        headers = list(map(lambda x: x.value, work_sheet[row_start]))
    last_non_empty_column_value = {}
    for row in islice(work_sheet.values, row_start, row_end if row_end else work_sheet.max_row):
        json_row = OrderedDict()
        i = 0
        column_end = column_end if column_end else len(headers)
        for header in headers:
            if header is None or (not custom_headers and (i < column_start or i > column_end)):
                i += 1
                continue
            header_name_v = header_name(header, config.get("key_replacements"))
            index = i + column_start if custom_headers else i
            value_v = value(row[index], config.get("value_replacements"), config.get("nulls_to_empty", True))
            if value_v is not None and value_v != "":
                last_non_empty_column_value[header_name_v] = value_v
            else:
                if config.get("repeat_column_value") and \
                        config.get("repeat_column_value").get(header_name_v) is not None:
                    value_v = last_non_empty_column_value[header_name_v]
            json_row[header_name_v] = value_v
            i += 1
        if not all_empty(json_row):
            json_list.append(json_row)
    return json_list


def get_all_constants():
    wb = load_workbook(filename='../../../stuff/excel_app/tools_default_data.xlsx', data_only=True)
    cc_config = {
        "key_replacements": {
            'average daily Solar rad [MJ/m^2] South vert. Wall': 'average_daily'
        }
    }
    td_config = {
        "key_replacements": {
            'building type': 'building_type',
            'age': 'building_year',
            'Wall thermal transmittance [W/(m2K)': 'wall_trans',
            'Roof thermal transmittance [W/(m2K)': 'roof_trans',
            'Floor thermal transmittance [W/(m2K)': 'floor_trans',
            'Windows thermal transmittance [W/(m2K)': 'windows_trans',
            'S disp/V': 'disp_v_ratio',
            'Swin/Sfloor': 'win_floor_ratio',
        },
    }
    vc_config = {
        "key_replacements": {
            'variation rate \\  year': 'variation_rate_per_year',
        },
        "value_replacements": {
            'electric energy [€/kWh]': 'electric_energy',
            'methane gas   [€/Smc]': 'methane_gas',
            'pellet [€/kg]': 'pellet',
        },
        "repeat_column_value": {
            "country": ""
        }
    }
    return {
        "city_climate": import_standard_table(wb['City Climate'], cc_config),
        "thermal_data": import_standard_table(wb['Thermal Data'], td_config),
        "variable_costs": import_standard_table(wb['Variable Costs'], vc_config),
    }
