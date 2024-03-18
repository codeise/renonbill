import unittest
from openpyxl import load_workbook
import json
from api.services.excel_import import import_standard_table


class TestExcelImport(unittest.TestCase):
    def setUp(self) -> None:
        self.wb = load_workbook(filename='tools_default_data.xlsx', data_only=True)
        # print(wb.sheetnames)

    def convert_and_print(self, sheet_name, config):
        sheet = self.wb[sheet_name]
        table = import_standard_table(sheet, config)
        j = json.dumps(table)
        print("-----------------------------------------------")
        print(j)

    def test_import_city_climate(self):
        config = {
            "key_replacements": {
                'average daily Solar rad [MJ/m^2] South vert. Wall': 'average_daily'
            }
        }
        self.convert_and_print('City Climate', config)

    def test_import_thermal_data(self):
        config = {
            "key_replacements": {
                'building type': 'building_type',
                'age': 'building_year',
                'Wall thermal transmittance [W/(m2K)': 'wall_trans',
                'Roof thermal transmittance [W/(m2K)': 'roof_trans',
                'Floor thermal transmittance [W/(m2K)': 'floor_trans',
                'Windows thermal transmittance [W/(m2K)': 'windows_trans',
                'S disp/V': 'disp_v_ratio',
                'Swin/Sfloor': 'win_floor_ratio',
            },
        }
        self.convert_and_print('Thermal Data', config)

    def test_import_variable_costs(self):
        config = {
            "key_replacements": {
                'variation rate \\  year': 'variation_rate_per_year',
            },
            "value_replacements": {
                'electric energy [€/kWh]': 'electric_energy',
                'methane gas   [€/Smc]': 'methane_gas',
                'pellet [€/kg]': 'pellet',
            },
            "repeat_column_value": {
                "country": ""
            }
        }
        self.convert_and_print('Variable Costs', config)

    def test_import_envelope_windows(self):
        sheet_name = 'Envelope & Windows'
        sheet = self.wb[sheet_name]
        config = {
            "key_replacements": {
                'EXTERNAL INSULATION': 'external_insulation',
                'thermal conductivity': 'thermal_conductivity',
                'material cost €/m2': 'material_cost',
                'installation cost € /m2': 'installation_cost',
            },
            "range": {
                "row_start": 1,
                "row_end": 4,
            }
        }
        table1 = import_standard_table(sheet, config)
        config = {
            "value_replacements": {
                'single grazed cost, €/m2 (all inclusive)': 'single_grazed_cost',
                'double glazed cost, €/m2 (all inclusive)': 'double_grazed_cost',
            },
            "range": {
                "row_start": 6,
                "row_end": 7,
            },
            "custom_headers": ["cost", None, None, "value"]
        }
        table2 = import_standard_table(sheet, config)

        complete = {"table1": table1, "table2": table2}
        j = json.dumps(complete)
        print("-----------------------------------------------")
        print(j)

    def test_import_heating_dhw(self):
        sheet_name = 'Heating & DHW'
        sheet = self.wb[sheet_name]
        config = {
            "key_replacements": {
                'HEATING (plant type)': 'heating',
                'cost €/kW': 'cost',
                'installation cost €/kW': 'installation_cost',
            },
            "value_replacements": {
                'solar heating (€/m2)': 'solar_heating',
            },
            "range": {
                "row_start": 1,
                "row_end": 13,
                "column_start": 1,
                "column_end": 4
            }
        }
        table1 = import_standard_table(sheet, config)

        config = {
            "custom_headers": ["plant_type", "efficiency", "cost", "installation_cost"],
            "value_replacements": {
                'electric boiler (€/l)': 'electric_boiler',
                'heat pump  (€/l)': 'heat_pump',
                'solar heater (€/m2)': 'solar_heater',
            },
            "range": {
                "row_start": 15,
                "row_end": 21,
                "column_start": 1,
                "column_end": 4
            }
        }
        table2 = import_standard_table(sheet, config)

        config = {
            "key_replacements": {
                'HEATING (plant type)': 'heating',
                'cost €/kW': 'cost',
                'installation cost €/kW': 'installation_cost',
            },
            "range": {
                "row_start": 1,
                "row_end": 7,
                "column_start": 5,
                "column_end": 8
            }
        }
        table3 = import_standard_table(sheet, config)

        config = {
            "key_replacements": {
                'regulation: mean efficiency': 'regulation_mean_efficiency',
                'distribution: mean efficiency': 'distribution_mean_efficiency',
            },
            "range": {
                "row_start": 1,
                "row_end": 2,
                "column_start": 9,
                "column_end": 10
            }
        }
        table4 = import_standard_table(sheet, config)

        complete = {"table1": table1, "table2": table2, "table3": table3, "table4": table4}
        j = json.dumps(complete)
        print("-----------------------------------------------")
        print(j)

    def test_import_other_thermal_data(self):
        config = {
            "custom_headers": ["description", "value", None, "note"],
        }
        self.convert_and_print('Other Thermal Data', config)

    def test_import_default_uncertainty(self):
        sheet_name = 'Default Uncertainty'
        sheet = self.wb[sheet_name]
        config = {
            "custom_headers": ["simplified_manager", "check", "conf_today", "conf_year"],
            "range": {
                "row_start": 3,
                "row_end": 9,
                "column_start": 1,
                "column_end": 4
            }
        }
        table1 = import_standard_table(sheet, config)

        config = {
            "custom_headers": ["advanced_manager", "desc", "check", "conf_today"],
            "range": {
                "row_start": 3,
                "row_end": 20,
                "column_start": 7,
                "column_end": 11
            },
            "repeat_column_value": {
                "advanced_manager": ""
            }
        }
        table2 = import_standard_table(sheet, config)

        config = {
            "custom_headers": ["advanced_manager1", "desc", "check", "conf_today", "conf_year"],
            "range": {
                "row_start": 3,
                "row_end": 28,
                "column_start": 12,
                "column_end": 17
            },
            "repeat_column_value": {
                "advanced_manager1": ""
            }
        }
        table3 = import_standard_table(sheet, config)

        config = {
            "range": {
                "row_start": 13,
                "row_end": 14,
                "column_start": 1,
                "column_end": 1
            }
        }
        table4 = import_standard_table(sheet, config)

        config = {
            "range": {
                "row_start": 13,
                "row_end": 14,
                "column_start": 4,
                "column_end": 4
            }
        }
        table5 = import_standard_table(sheet, config)

        complete = {"table1": table1, "table2": table2, "table3": table3, "table4": table4, "table5": table5}
        j = json.dumps(complete)
        print("-----------------------------------------------")
        print(j)

    def test_import_uncertain_variables(self):
        sheet_name = 'uncertain variables'
        sheet = self.wb[sheet_name]
        config = {
            "custom_headers": ["geometric", "value"],
            "range": {
                "row_start": 4,
                "row_end": 7,
                "column_start": 2,
                "column_end": 3
            }
        }
        table1 = import_standard_table(sheet, config)

        config = {
            "custom_headers": ["thermophysical", "value"],
            "range": {
                "row_start": 4,
                "row_end": 15,
                "column_start": 4,
                "column_end": 5
            }
        }
        table2 = import_standard_table(sheet, config)

        config = {
            "custom_headers": ["heating_equipment", "value"],
            "range": {
                "row_start": 4,
                "row_end": 16,
                "column_start": 6,
                "column_end": 7
            }
        }
        table3 = import_standard_table(sheet, config)

        config = {
            "custom_headers": ["external_environment", "value"],
            "range": {
                "row_start": 4,
                "row_end": 9,
                "column_start": 8,
                "column_end": 9
            }
        }
        table4 = import_standard_table(sheet, config)

        config = {
            "custom_headers": ["economic", "value"],
            "range": {
                "row_start": 4,
                "row_end": 8,
                "column_start": 10,
                "column_end": 11
            }
        }
        table5 = import_standard_table(sheet, config)

        complete = {"table1": table1, "table2": table2, "table3": table3, "table4": table4, "table5": table5}
        j = json.dumps(complete)
        print("-----------------------------------------------")
        print(j)
